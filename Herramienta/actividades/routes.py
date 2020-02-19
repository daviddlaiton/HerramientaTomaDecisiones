import os
import statistics 
from flask import render_template, url_for, flash, redirect, request, Blueprint, abort, current_app, send_file
from flask_wtf import FlaskForm
from wtforms import FieldList, FormField, SubmitField
from flask_login import current_user, login_required
from Herramienta.models import Usuario, Curso, Semestre, Actividad, Punto, Inciso, Criterio, Subcriterio, Variacion, Grupo, Estudiante, Calificacion, ListaUsuariosSemestreCurso
from Herramienta import db, bcrypt
from Herramienta.actividades.forms import CrearActividadArchivoForm, EliminarActividad, DescargarActividad, CrearPunto, CambiarEstadoActividad, EnviarReportes, IntegranteForm, EscogerGrupoParaCalificar, EliminarGrupo, DescargarFormatoActividadForm, GenerarReporte
from openpyxl import load_workbook, Workbook
from Herramienta.actividades.utils import send_reports, create_pdf

actividades = Blueprint("actividades", __name__)

@actividades.route("/actividades/<int:curso_id>/<int:actividad_id>", methods=["GET", "POST"])
@login_required
def ver_actividad(actividad_id,curso_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    actividad = Actividad.query.get_or_404(actividad_id)
    return render_template("actividades/ver_actividad.html", title="Ver actividad", actividad=actividad, curso_id=curso_id)

@actividades.route("/actividades/<int:curso_id>/<int:actividad_id>/elegirGrupoCalificar", methods=["GET", "POST"])
@login_required
def elegir_grupo_calificar_actividad(actividad_id,curso_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    actividad = Actividad.query.get_or_404(actividad_id)
    grupos = [(g.id, g.estudiantes) for g in Grupo.query.filter_by(actividad_id=actividad.id, calificaciones=None).all()]
    form = EscogerGrupoParaCalificar(request.form)
    form.grupo.choices = grupos
    if form.validate_on_submit():
        return redirect(url_for("actividades.calificar_actividad", curso_id=curso_id, actividad_id=actividad_id, grupo_id=form.grupo.data))
    return render_template("actividades/elegir_grupo_calificar.html", title="Elegir grupo para calificar actividad", actividad=actividad, curso_id=curso_id,form=form, grupos=grupos)

@actividades.route("/actividades/<int:curso_id>/<int:actividad_id>/calificar/<int:grupo_id>", methods=["GET", "POST"])
@login_required
def calificar_actividad(actividad_id,curso_id,grupo_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        habilitado = ListaUsuariosSemestreCurso.query.filter_by(usuario_id=user_id, curso_id=curso_id).first()
        if not habilitado:
            abort(403)
    actividad = Actividad.query.get_or_404(actividad_id)
    grupo = Grupo.query.get_or_404(grupo_id)
    puntaje = 0
    numSubcriterios = 0
    puntos = []
    calificacionesExistentes = []
    for punto in actividad.puntos:
        incisos = []
        for inciso in punto.incisos:
            criterios = []
            for criterio in inciso.criterios:
                subcriterios = []
                for subcriterio in criterio.subcriterios:
                    variaciones = []
                    for variacion in subcriterio.variaciones:
                        calificacionExistente = Calificacion.query.filter_by(grupo_id=grupo_id, variacion_id=variacion.id).first()
                        if calificacionExistente is not None:
                            calificacionesExistentes.append(str(variacion.id) + ":" + str(subcriterio.id))
                        variacionJSON = {
                            "id" : variacion.id,
                            "puntaje" : variacion.puntaje,
                            "esOtro" : variacion.esOtro
                        }
                        variaciones.append(variacionJSON)
                    puntaje = puntaje + subcriterio.maximoPuntaje
                    if subcriterio.maximoPuntaje > 0:
                        numSubcriterios = numSubcriterios + 1
                    subcriterioJSON = {
                        "id" : subcriterio.id,
                        "variaciones" : variaciones
                    }
                    subcriterios.append(subcriterioJSON)
                criterioJSON = {
                    "id" : criterio.id,
                    "subcriterios" : subcriterios
                }
                criterios.append(criterioJSON)
            incisoJSON = {
                "id" : inciso.id,
                "criterios" : criterios
            }
            incisos.append(incisoJSON)
        puntoJSON = {
            "id" : punto.id,
            "incisos" : incisos
        }
        puntos.append(puntoJSON)
    actividadToJson = {
        "id" : actividad.id,
        "puntos" : puntos,
        "puntaje" : puntaje,
        "numSubcriterios" : numSubcriterios,
        "calificacionesExistentes" : calificacionesExistentes
    }
    return render_template("actividades/calificar_actividad.html", title="Calificar actividad", actividad=actividad, curso_id=curso_id, actividadJSON = actividadToJson, grupo=grupo)

@actividades.route("/actividades/<int:curso_id>/<int:actividad_id>/<int:semestre_id>/eliminar", methods=["GET", "POST"])
@login_required
def eliminar_actividad(actividad_id, curso_id,semestre_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    actividad = Actividad.query.get_or_404(actividad_id)
    grupos = Grupo.query.filter_by(actividad_id=actividad.id).all()
    if user.rol_id == 1:
        abort(403)
    form = EliminarActividad()
    if form.validate_on_submit():
        for grupo in grupos:
            listaCalificaciones = Calificacion.query.filter_by(grupo_id=grupo.id).all()
            for calificacion in listaCalificaciones:
                db.session.delete(calificacion)
                db.session.commit()
            db.session.delete(grupo)
            db.session.commit()
        eliminarActividad(actividad_id)
        flash(f"Actividad eliminada exitosamente", "success")
        return redirect(url_for("actividades.verActividades_semestre", curso_id=curso_id, semestre_id=semestre_id))
    return render_template("actividades/eliminar_actividad.html", title="Eliminar actividad", curso_id=curso_id, actividad_id=actividad_id, semestre_id=semestre_id,form=form)

@actividades.route("/actividades/<int:curso_id>/<int:actividad_id>/<int:semestre_id>/cambiarEstado", methods=["GET", "POST"])
@login_required
def cambiarEstado_actividad(actividad_id,curso_id,semestre_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    actividad = Actividad.query.get_or_404(actividad_id)
    actividad.habilitada  = not actividad.habilitada
    db.session.commit()
    return redirect(url_for("actividades.verActividades_semestre", curso_id=curso_id, semestre_id=semestre_id))

@actividades.route("/actividades/<int:curso_id>/<int:semestre_id>/crearActividad")
@login_required
def crear_actividad(curso_id, semestre_id):
    return render_template("actividades/crear_actividad.html", title="Crear actividad", curso_id=curso_id, semestre_id=semestre_id)

@actividades.route("/actividades/<int:curso_id>/<int:semestre_id>/crearActividadArchivo", methods=["GET", "POST"])
@login_required
def crear_actividadArchivo(curso_id, semestre_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    form = CrearActividadArchivoForm()
    if form.validate_on_submit():
        if form.archivo.data:
            if request.method == 'POST':
                f = request.files['archivo']
                f.save(os.path.join(current_app.root_path, 'static/files', "Actividad.xlsx"))
                analisis = analizarArchivo(curso_id,semestre_id)
                if analisis is None:
                    flash(f"Actividad creada exitosamente", "success")
                elif analisis == "nombre":
                    flash(f"El archivo no pudo ser procesado porque ya existe una actividad con ese nombre.", "danger")
                elif analisis == "porcentaje":
                    flash(f"El archivo no pudo ser procesado porque el porcentaje no es un numero.", "danger")
                elif analisis == "porcentajeMayorACero":
                    flash("El archivo no pudo ser procesado porque el valor de porcentaje es mayor a 1", "danger")
                elif analisis == "integrantes":
                    flash("El archivo no pudo ser procesado porque el número de integrantes es menor o igual a 0.", "danger")
                elif analisis.startswith("formato"):
                    lugar = analisis.split(":")
                    flash("El archivo no pudo ser procesado porque hay un error en la fila " + lugar[1] + " columna " +lugar[2], "danger")
                elif analisis.startswith("2formato"):
                    lugar = analisis.split(":")
                    flash("El archivo no pudo ser procesado porque hay un error en la fila " + lugar[1], "danger")
                return redirect(url_for("actividades.verActividades_semestre", curso_id=curso_id, semestre_id=semestre_id))
    return render_template("actividades/crear_actividadArchivo.html", title="Crear actividad desde archivo", curso_id=curso_id, form=form, semestre_id=semestre_id)


def analizarArchivo(curso_id, semestre_id):
    tipoError = None
    archivoExcel = load_workbook(current_app.root_path + '/static/files/Actividad.xlsx')
    hoja = archivoExcel.active
    nombre = hoja["B1"].value
    if Actividad.query.filter_by(nombre=nombre, semestre_id=semestre_id, curso_id=curso_id).first() is not None:
        tipoError = "nombre"
        return tipoError
    porcentaje = 0
    try:
        porcentaje = float(hoja["B2"].value)
    except ValueError:
        tipoError = "porcentaje"
        return tipoError
    if porcentaje > 1:
        tipoError = "porcentajeMayorACero"
        return tipoError
    numeroIntegrantes = int(hoja["B3"].value) 
    if numeroIntegrantes <= 0:
        tipoError = "integrantes"
        return tipoError
    actividad = Actividad(nombre=nombre, porcentaje=porcentaje, habilitada=False, semestre_id=semestre_id, curso_id=curso_id, numeroIntegrantes=numeroIntegrantes, numGrupos=0, numEstCalificados=0, promedio=0, desvEst=0)
    db.session.add(actividad)
    db.session.commit()
    #Siempre se debe comenzar ahí, el formato se tiene que respetar.
    fila = 5
    columna = 2
    final = False
    celdaPunto = hoja.cell(row=fila, column=columna).value
    if celdaPunto is None:
        eliminarActividad(actividad.id)
        tipoError = "formato:" + str(fila) + ":" + str(columna)
        return tipoError

    while not final:
        if celdaPunto is not None:
            punto = Punto(nombre=celdaPunto, actividad_id=actividad.id, puntajePosible=0)
            db.session.add(punto)
            db.session.commit()
            finalPunto = False
            fila = fila + 1
            columna = columna + 1
            celdaInciso = hoja.cell(row=fila, column=columna).value
            #-------------------------------------------------------------------
            while not finalPunto:
                if celdaInciso is None:
                    eliminarActividad(actividad.id)
                    tipoError = "formato:" + str(fila) + ":" + str(columna)
                    return tipoError
                inciso = Inciso(nombre = celdaInciso, puntajePosible=0, punto_id=punto.id)
                db.session.add(inciso)
                db.session.commit()
                finalInciso = False
                fila = fila + 1
                columna = columna + 1
                celdaCriterio = hoja.cell(row=fila, column=columna).value
                #-------------------------------------------------------------------
                while not finalInciso:
                    if celdaCriterio is None:
                        eliminarActividad(actividad.id)
                        tipoError = "formato:" + str(fila) + ":" + str(columna)
                        return tipoError
                    criterio = Criterio(nombre = celdaCriterio, puntajePosible=0, inciso_id=inciso.id)
                    db.session.add(criterio)
                    db.session.commit()
                    finalCriterio = False
                    fila = fila + 1
                    columna = columna + 1
                    celdaSubriterio = hoja.cell(row=fila, column=columna).value
                    #-------------------------------------------------------------------
                    while not finalCriterio:
                        if celdaSubriterio is None:
                            eliminarActividad(actividad.id)
                            tipoError = "formato:" + str(fila) + ":" + str(columna)
                            return tipoError                        
                        puntajeMinimo = float(hoja.cell(row=fila, column=7).value)
                        puntajeMaximo = float(hoja.cell(row=fila, column=8).value)
                        subcriterio = Subcriterio(nombre = celdaSubriterio, maximoPuntaje=puntajeMaximo, minimoPuntaje=puntajeMinimo, criterio_id=criterio.id)
                        db.session.add(subcriterio)
                        db.session.commit()
                        finalSubcriterio = False
                        fila = fila + 1
                        columna = columna + 1
                        celdaVariacion = hoja.cell(row=fila, column=columna).value
                        #-------------------------------------------------------------------
                        while not finalSubcriterio:
                            maximoVeces = int(hoja.cell(row=fila, column=9).value)
                            puntaje = float(hoja.cell(row=fila, column=8).value)
                            esPenalizacion = False
                            if puntaje < 0:
                                esPenalizacion = True
                            variacion = Variacion(descripcion = celdaVariacion, puntaje=puntaje, esPenalizacion=esPenalizacion, subcriterio_id=subcriterio.id, esOtro=False, maximoVeces=maximoVeces)
                            db.session.add(variacion)
                            db.session.commit()
                            subcriterio.variaciones.append(variacion)
                            db.session.commit()
                            fila = fila + 1
                            celdaVariacion = hoja.cell(row=fila, column=columna).value
                            if celdaVariacion is None:
                                finalSubcriterio = True
                                variacion = Variacion(descripcion = "No realizó nada", puntaje=0, esPenalizacion=False, subcriterio_id=subcriterio.id, esOtro=True, maximoVeces=1)
                                db.session.add(variacion)
                                db.session.commit()
                                subcriterio.variaciones.append(variacion)
                                db.session.commit()
                        #-------------------------------------------------------------------
                        columna = columna - 1
                        celdaSubriterio = hoja.cell(row=fila, column=columna).value
                        if celdaSubriterio is None:
                            finalCriterio = True
                        criterio.puntajePosible = criterio.puntajePosible + subcriterio.maximoPuntaje
                        criterio.subcriterios.append(subcriterio)
                        db.session.commit()
                    #-------------------------------------------------------------------
                    columna = columna - 1
                    celdaCriterio = hoja.cell(row=fila, column=columna).value
                    if celdaCriterio is None:
                        finalInciso = True
                    inciso.puntajePosible = inciso.puntajePosible + criterio.puntajePosible
                    inciso.criterios.append(criterio)
                    db.session.commit()
                #-------------------------------------------------------------------
                columna = columna - 1
                celdaInciso = hoja.cell(row=fila, column=columna).value
                if celdaInciso is None:
                    finalPunto = True
                punto.puntajePosible = punto.puntajePosible + inciso.puntajePosible
                punto.incisos.append(inciso)
                db.session.commit()
            #-------------------------------------------------------------------
            columna = columna - 1
            celdaPunto = hoja.cell(row=fila, column=columna).value
            actividad.puntos.append(punto)
            db.session.commit()
        else:
            finalVerdadero = False
            fila = fila + 1
            columna = 1
            while not finalVerdadero:
                celda = hoja.cell(row=fila, column=columna).value
                if celda is not None:
                    eliminarActividad(actividad.id)
                    tipoError = "2formato:" + str(fila-1)
                    return tipoError
                columna = columna + 1
                if columna >= 6:
                    finalVerdadero = True
            final = True          
@actividades.route("/actividades/<int:curso_id>/crearActividadWeb", methods=["GET", "POST"])
@login_required
def crear_actividadWeb(curso_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    return render_template("actividades/crear_actividadWeb.html", title="Crear actividad Web", curso_id=curso_id)

@actividades.route("/actividades/<int:curso_id>/<int:actividad_id>/descargar", methods=["GET","POST"])
@login_required
def descargar_actividad(actividad_id,curso_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    actividad = Actividad.query.get_or_404(actividad_id)
    form = DescargarActividad()
    wb = Workbook()
    dest_filename = 'Herramienta/static/files/ExportarActividad.xlsx'
    hoja = wb.active
    hoja.cell(column=1, row=1, value="Nombre:")
    hoja.cell(column=2, row=1, value=actividad.nombre)
    hoja.cell(column=1, row=2, value="Porcentaje sobre la nota:")
    hoja.cell(column=2, row=2, value=actividad.porcentaje)
    hoja.cell(column=1, row=3, value="Integrantes por grupo:")
    hoja.cell(column=2, row=3, value=actividad.numeroIntegrantes)


    hoja.cell(column=1, row=4, value="ID")
    hoja.cell(column=2, row=4, value="Punto")
    hoja.cell(column=3, row=4, value="Inciso")
    hoja.cell(column=4, row=4, value="Criterio")
    hoja.cell(column=5, row=4, value="Subcriterio")
    hoja.cell(column=6, row=4, value="Variación/Penalización")
    hoja.cell(column=7, row=4, value="Puntaje minimo")
    hoja.cell(column=8, row=4, value="Puntaje")
    hoja.cell(column=9, row=4, value="Máximo veces")

    
    #------------------------------------------------------------------
    fila = 5
    idPunto = 1
    idInciso = 1
    idCriterio = 1
    idSubcriterio = 1
    idVariacion = 1

    for punto in actividad.puntos:
        hoja.cell(column=1, row=fila, value=idPunto)
        hoja.cell(column=2, row=fila, value=punto.nombre)
        idPunto = idPunto + 1
        fila = fila + 1
        for inciso in punto.incisos:
            hoja.cell(column=1, row=fila, value=idInciso)
            hoja.cell(column=3, row=fila, value=inciso.nombre)
            idInciso = idInciso + 1
            fila = fila + 1
            for criterio in inciso.criterios:
                hoja.cell(column=1, row=fila, value=idCriterio)
                hoja.cell(column=4, row=fila, value=criterio.nombre)
                idCriterio = idCriterio + 1
                fila = fila + 1
                for subcriterio in criterio.subcriterios:
                    hoja.cell(column=1, row=fila, value=idSubcriterio)
                    hoja.cell(column=5, row=fila, value=subcriterio.nombre)
                    hoja.cell(column=7, row=fila, value=subcriterio.minimoPuntaje)
                    hoja.cell(column=8, row=fila, value=subcriterio.maximoPuntaje)
                    idSubcriterio = idSubcriterio + 1
                    fila = fila + 1
                    for variacion in subcriterio.variaciones:
                        hoja.cell(column=1, row=fila, value=idVariacion)
                        hoja.cell(column=6, row=fila, value=variacion.descripcion)
                        hoja.cell(column=8, row=fila, value=variacion.puntaje)
                        hoja.cell(column=9, row=fila, value=variacion.maximoVeces)
                        idVariacion = idVariacion + 1
                        fila = fila + 1
    hoja.column_dimensions['E'].width = 10
    hoja.column_dimensions['F'].width = 200
    wb.save(filename = dest_filename)
    if form.validate_on_submit():
        return send_file('static/files/ExportarActividad.xlsx',
                     mimetype='text/xlsx',
                     attachment_filename='ExportarActividad.xlsx',
                     as_attachment=True)
    return render_template("actividades/descargar_actividad.html", title="Descargar actividad Web", actividad_id=actividad_id, curso_id=curso_id, form=form)

def eliminarActividad(actividad_id):
    actividad = Actividad.query.get_or_404(actividad_id)
    for punto in actividad.puntos:
        #-------------------------------------------------------------------
        for inciso in punto.incisos:
            #-------------------------------------------------------------------
            for criterio in inciso.criterios:
                #-------------------------------------------------------------------
                for subcriterio in criterio.subcriterios:
                    #-------------------------------------------------------------------
                    for variacion in subcriterio.variaciones:
                        db.session.delete(variacion)
                        db.session.commit()
                    #-------------------------------------------------------------------
                    db.session.delete(subcriterio)
                    db.session.commit()
                #-------------------------------------------------------------------
                db.session.delete(criterio)
                db.session.commit()
            #-------------------------------------------------------------------
            db.session.delete(inciso)
            db.session.commit()
        #-------------------------------------------------------------------
        db.session.delete(punto)
        db.session.commit()

    db.session.delete(actividad)
    db.session.commit()

@actividades.route("/actividades/<int:curso_id>/<int:actividad_id>/crearPunto", methods=["GET","POST"])
@login_required
def crear_punto(curso_id,actividad_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    actividad = Actividad.query.get_or_404(actividad_id)
    # form = EliminarActividad()
    # punto = Punto(nombre=, actividad_id=actividad.id, puntajePosible=0)

@actividades.route("/actividades/<int:curso_id>/<int:actividad_id>/enviarInformes", methods=["GET","POST"])
@login_required
def enviar_informes(curso_id,actividad_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    actividad = Actividad.query.get_or_404(actividad_id)
    form = EnviarReportes()
    if form.validate_on_submit():
        send_reports(actividad)
        flash(f"Informes enviados exitosamente", "success")
        return render_template("actividades/ver_actividad.html", title="Ver actividad", actividad=actividad, curso_id=curso_id)
    return render_template("actividades/enviarReportes.html", title="Enviar reportes", actividad_id=actividad_id, curso_id=curso_id, form=form)

@actividades.route("/actividades/<int:curso_id>/<int:actividad_id>/generarInforme/<int:grupo_id>", methods=["GET","POST"])
@login_required
def generar_informe(curso_id,actividad_id,grupo_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    actividad = Actividad.query.get_or_404(actividad_id)
    grupo = Grupo.query.get_or_404(grupo_id)
    form = GenerarReporte()
    if form.validate_on_submit():
        create_pdf(actividad,grupo)
        return send_file('actividades/files/reporte.pdf',
                        mimetype='application/pdf',
                        attachment_filename='reporte.pdf',
                        as_attachment=True)
    return render_template("actividades/generarReporte.html", title="Generar reporte", actividad_id=actividad_id, curso_id=curso_id, form=form, grupo=grupo)
    
@actividades.route("/actividades/<int:curso_id>/<int:actividad_id>/verGrupos", methods=["GET", "POST"])
@login_required
def ver_grupos_actividad(actividad_id,curso_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    actividad = Actividad.query.get_or_404(actividad_id)
    semestre = Semestre.query.get_or_404(actividad.semestre_id)
    curso = Curso.query.get_or_404(curso_id)
    grupos = []
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        grupos = Grupo.query.filter_by(actividad_id=actividad_id,usuario_id=user_id).all()
    else:
        grupos = Grupo.query.filter_by(actividad_id=actividad_id).all()
    return render_template("actividades/ver_grupos_actividad.html", title="Ver grupos", actividad=actividad, curso=curso, grupos=grupos, semestre=semestre)

@actividades.route("/actividades/<int:curso_id>/<int:actividad_id>/crearGrupo/<int:numero_integrantes>", methods=["GET", "POST"])
@login_required
def crear_grupo_actividad(actividad_id,curso_id, numero_integrantes):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    actividad = Actividad.query.get_or_404(actividad_id)
    estudiantes = Semestre.query.get_or_404(actividad.semestre_id).estudiantes
    estudiantesEnGrupo = []
    estudiantesJSON = []
    for estudiante in estudiantes:
        estudianteAnadir = {
            "codigo" : str(estudiante.codigo),
            "login" : estudiante.login,
            "nombres" : estudiante.nombre,
            "apellidos" : estudiante.apellido
        }
        estudiantesJSON.append(estudianteAnadir)
    for grupo in actividad.grupos:
        estudiantesGrupo = []
        for estudiante in grupo.estudiantes:
            estudiantesEnGrupo.append(estudiante.codigo)
        estudiantesEnGrupo.append(estudiantesGrupo)

    return render_template("actividades/crear_grupo_actividad.html", title="Crear grupos", actividad=actividad, curso_id=curso_id, numero_integrantes=numero_integrantes, estudiantesJSON = estudiantesJSON, estudiantesEnGrupo=estudiantesEnGrupo)

@actividades.route("/actividades/<int:curso_id>/<int:actividad_id>/grupoCreado/<integrantesSeleccionados>", methods=["GET", "POST"])
@login_required
def grupo_creado_actividad(actividad_id,curso_id,integrantesSeleccionados):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    actividad = Actividad.query.get_or_404(actividad_id)
    integrantes = []
    actividad.numGrupos = actividad.numGrupos + 1
    numeroGrupo = actividad.numGrupos
    codigosIntegrantes = integrantesSeleccionados.split(":")
    for codigoIntegrante in codigosIntegrantes:
        if codigoIntegrante is not "":
            integrante = Estudiante.query.filter_by(codigo=codigoIntegrante).first()
            integrantes.append(integrante)
    grupo = Grupo(actividad_id=actividad_id, estudiantes=integrantes, numero=numeroGrupo, usuario_id=current_user.get_id() , creador=current_user.login, nota = 0, estadoCalifacion="SinEmpezar")
    db.session.add(grupo)
    db.session.commit()
    flash(f"Grupo creado exitosamente", "success")
    return redirect(url_for("actividades.ver_grupos_actividad", curso_id=curso_id, actividad_id=actividad_id))

@actividades.route("/actividades/<int:curso_id>/<int:semestre_id>/", methods=["GET", "POST"])
@login_required
def verActividades_semestre(semestre_id,curso_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    curso = Curso.query.get_or_404(curso_id)
    semestre = Semestre.query.get_or_404(semestre_id)
    actividades = []
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        habilitado = ListaUsuariosSemestreCurso.query.filter_by(usuario_id=user_id, curso_id=curso_id,semestre_id=semestre_id).first()
        if not habilitado:
            abort(403)
        actividades = Actividad.query.filter_by(semestre_id=semestre_id, curso_id=curso_id, habilitada=True).all()
    else:
        actividades = Actividad.query.filter_by(semestre_id=semestre_id, curso_id=curso_id).all()
    return render_template("actividades/ver_actividades_semestre.html", title="Ver actividades", actividades=actividades, curso=curso, semestre=semestre)

@actividades.route("/actividades/<int:curso_id>/<int:actividad_id>/eliminarGrupo/<int:grupo_id>", methods=["GET", "POST"])
@login_required
def eliminar_grupo_semestre(actividad_id,curso_id,grupo_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    curso = Curso.query.get_or_404(curso_id)
    actividad = Actividad.query.get_or_404(actividad_id)
    grupo = Grupo.query.get_or_404(grupo_id)
    form = EliminarGrupo()
    if form.validate_on_submit():
        listaCalificaciones = Calificacion.query.filter_by(grupo_id=grupo.id).all()
        for calificacion in listaCalificaciones:
            db.session.delete(calificacion)
            db.session.commit()
        actividad.grupos.remove(grupo)
        db.session.delete(grupo)
        db.session.commit()
        flash(f"Grupo eliminado exitosamente", "success")
        return redirect(url_for("actividades.ver_grupos_actividad", curso_id=curso_id, actividad_id=actividad.id))
    return render_template("actividades/eliminar_grupo.html", title="Eliminar grupo", actividad=actividad, curso=curso,grupo=grupo, form=form)

@actividades.route("/actividades/<int:semestre_id>/<int:curso_id>/descargarFormatoLista", methods=["GET", "POST"])
@login_required
def descargarFormatoActividad(semestre_id,curso_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    form = DescargarFormatoActividadForm()
    if form.validate_on_submit():
        return send_file('static/files/FormatoEjemploTarea.xlsx',
                     mimetype='text/xlsx',
                     attachment_filename='FormatoEjemploTarea.xlsx',
                     as_attachment=True)
    return render_template("actividades/descargarFormatoActividades.html", title="Descargar formato actividades", curso_id=curso_id, form=form, semestre_id=semestre_id)


@actividades.route("/actividades/<int:actividad_id>/<int:curso_id>/<int:grupo_id>/guardarNotas/<variaciones>/<estado>/<nota>/<puntaje>", methods=["GET", "POST"])
@login_required
def guardarNotas(actividad_id,curso_id, grupo_id, variaciones, estado, nota,puntaje):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    grupo = Grupo.query.get_or_404(grupo_id)
    actividad = Actividad.query.get_or_404(actividad_id)
    if estado == "Finalizado" and grupo.estadoCalifacion != "Finalizado":
        actividad.numEstCalificados = actividad.numEstCalificados + 1
    elif estado != "Finalizado" and grupo.estadoCalifacion == "Finalizado":
        actividad.numEstCalificados = actividad.numEstCalificados - 1
    grupo.estadoCalifacion = estado
    grupo.nota = ((float(nota)/float(puntaje))*5)
    db.session.commit()
    listaGrupos = Grupo.query.filter_by(actividad_id=actividad_id, estadoCalifacion="Finalizado").all()
    listaNotas = []
    for grupoL in listaGrupos:
        listaNotas.append(grupoL.nota) 
    if len(listaNotas) >1:
        actividad.desvEst = statistics.stdev(listaNotas)
    if len(listaNotas) > 0:
        actividad.promedio = statistics.mean(listaNotas)
    listaVariaciones = variaciones.split(":")
    listaVariacionesInt = []
    for variacion in listaVariaciones:
        if variacion != "":
            listaVariacionesInt.append(int(variacion))
    for punto in actividad.puntos:
        #-------------------------------------------------------------------
        for inciso in punto.incisos:
            #-------------------------------------------------------------------
            for criterio in inciso.criterios:
                #-------------------------------------------------------------------
                for subcriterio in criterio.subcriterios:
                    #-------------------------------------------------------------------
                    for variacion in subcriterio.variaciones:
                        calificacionExistente = Calificacion.query.filter_by(grupo_id=grupo_id, variacion_id=variacion.id).first()
                        if variacion.id in listaVariacionesInt:
                            if calificacionExistente is None:
                                nuevaCalificacion = Calificacion(grupo_id=grupo_id, variacion_id=variacion.id, descripcion=variacion.descripcion, cantidadVeces=1, puntaje=variacion.puntaje)
                                db.session.add(nuevaCalificacion)
                                listaVariacionesInt.remove(variacion.id)
                        else:
                            if calificacionExistente is not None:
                                db.session.delete(calificacionExistente)
                        db.session.commit()
    return redirect(url_for('actividades.ver_grupos_actividad', curso_id=curso_id, actividad_id=actividad_id))

@actividades.route("/actividades/<int:curso_id>/<int:actividad_id>/descargarNotas", methods=["GET","POST"])
@login_required
def descargar_notas(actividad_id,curso_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    actividad = Actividad.query.get_or_404(actividad_id)
    form = DescargarActividad()
    wb = Workbook()
    dest_filename = 'Herramienta/static/files/Notas.xlsx'
    hoja = wb.active
    hoja.cell(column=1, row=2, value="Código")
    hoja.cell(column=2, row=2, value="Login")
    hoja.cell(column=3, row=2, value="Apellidos")
    hoja.cell(column=4, row=2, value="Nombre")
    hoja.cell(column=5, row=2, value="Subcriterio")
    
    #------------------------------------------------------------------
    columna = 6

    for punto in actividad.puntos:
        hoja.cell(column=columna, row=1, value=punto.nombre)
        hoja.cell(column=columna, row=2, value=punto.puntajePosible)
        columna = columna + 1
        for inciso in punto.incisos:
            hoja.cell(column=columna, row=1, value=inciso.nombre)
            hoja.cell(column=columna, row=2, value=inciso.puntajePosible)
            columna = columna + 1
            for criterio in inciso.criterios:
                hoja.cell(column=columna, row=1, value=criterio.nombre)
                hoja.cell(column=columna, row=2, value=criterio.puntajePosible)
                columna = columna + 1
                for subcriterio in criterio.subcriterios:
                    hoja.cell(column=columna, row=1, value=subcriterio.nombre)
                    hoja.cell(column=columna, row=2, value=subcriterio.maximoPuntaje)
                    columna = columna + 1
                    for variacion in subcriterio.variaciones:
                        hoja.cell(column=columna, row=1, value=variacion.descripcion)
                        hoja.cell(column=columna, row=2, value=variacion.puntaje)
                        columna = columna + 1
    wb.save(filename = dest_filename)
    if form.validate_on_submit():
        return send_file('static/files/Notas.xlsx',
                     mimetype='text/xlsx',
                     attachment_filename='Notas.xlsx',
                     as_attachment=True)
    return render_template("actividades/descargar_notas.html", title="Descargar actividad Web", actividad_id=actividad_id, curso_id=curso_id, form=form)