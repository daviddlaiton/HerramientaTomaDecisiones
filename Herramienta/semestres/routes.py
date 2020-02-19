import os
from flask import render_template, url_for, flash, redirect, request, Blueprint, abort, current_app, send_file
from flask_login import current_user, login_required
from Herramienta.models import Usuario, Semestre, Curso, Estudiante, SemestreCursoHabilitados, Actividad, Calificacion, Grupo
from Herramienta import db, bcrypt
from Herramienta.semestres.forms import (CrearSemestreForm, EditarNombreSemestreForm,
                                         AgregarCursoASemestreForm, EliminarCursoASemestreForm, EliminarSemestreForm, CargarListaEstudiantesForm, DescargarListaEstudiantesForm, DescargarFormatoListaEstudiantesForm, EstudianteForm, EliminarEstudianteForm)
from openpyxl import load_workbook, Workbook

semestres = Blueprint("semestres", __name__)

@semestres.route("/semestres")
@login_required
def get_semestres():
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    page = request.args.get("page", 1, type=int)
    semestres = Semestre.query.order_by(
        Semestre.id.desc()).paginate(page=page, per_page=5)
    return render_template("semestres/semestres.html", title="Semestres", semestres=semestres)


@semestres.route("/crear_semestre", methods=["GET", "POST"])
@login_required
def crear_semestre():
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    form = CrearSemestreForm()
    if form.validate_on_submit():
        semestre = Semestre(nombre=form.nombre.data)
        db.session.add(semestre)
        db.session.commit()
        flash(f"Semestre creado exitosamente", "success")
        return redirect(url_for("semestres.get_semestres"))
    return render_template("semestres/crear_semestre.html", title="Crear semestre", form=form)


@semestres.route("/semestres/<int:semestre_id>", methods=["GET", "POST"])
@login_required
def ver_semestre(semestre_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    semestre = Semestre.query.get_or_404(semestre_id)
    return render_template("semestres/ver_semestre.html", title="Editar semestre", semestre=semestre)


@semestres.route("/semestres/<int:semestre_id>/editarNombre", methods=["GET", "POST"])
@login_required
def editarNombre_semestre(semestre_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    semestre = Semestre.query.get_or_404(semestre_id)
    form = EditarNombreSemestreForm()
    if form.validate_on_submit():
        semestre.nombre = form.nombre.data
        db.session.commit()
        flash(f"Semestre actualizado exitosamente", "success")
        return redirect(url_for("semestre.ver_semestre"))
    elif request.method == "GET":
        form.nombre.data = semestre.nombre
    return render_template("semestres/editar_nombre_semestre.html", title="Cambiar nombre semestre", form=form, semestre=semestre)


@semestres.route("/semestres/<int:semestre_id>/agregarCurso", methods=["GET", "POST"])
@login_required
def agregarCurso_semestre(semestre_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    semestre = Semestre.query.get_or_404(semestre_id)
    cursosActuales = semestre.cursos
    cursos = [(c.id, c.nombre)
              for c in Curso.query.all() if c not in cursosActuales]
    form = AgregarCursoASemestreForm(request.form)
    form.curso.choices = cursos
    if form.validate_on_submit():
        cursoAnadir = Curso.query.filter_by(id=form.curso.data).first()
        semestre.cursos.append(cursoAnadir)
        semestreCurso = SemestreCursoHabilitados(semestre_id=semestre_id,curso_id=cursoAnadir.id,habilitado=True)
        db.session.add(semestreCurso)
        db.session.commit()
        flash(f"Curso añadido exitosamente", "success")
        return redirect(url_for("semestres.ver_semestre", semestre_id=semestre.id))
    return render_template("semestres/anadirCurso_semestre.html", title="Anadir curso a semestre", form=form, semestre=semestre, cursos=cursos)


@semestres.route("/semestres/<int:semestre_id>/eliminarCurso/<int:curso_id>", methods=["GET", "POST"])
@login_required
def eliminarCurso_semestre(semestre_id, curso_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    semestre = Semestre.query.get_or_404(semestre_id)
    curso = Curso.query.get_or_404(curso_id)
    form = EliminarCursoASemestreForm()
    if form.validate_on_submit():
        semestre.cursos.remove(curso)
        semestreCurso = SemestreCursoHabilitados.query.filter_by(semestre_id=semestre_id, curso_id=curso_id).first()
        db.session.delete(semestreCurso)
        db.session.commit()
        flash(f"Curso eliminado exitosamente", "success")
        return redirect(url_for("semestres.ver_semestre", semestre_id=semestre.id))
    return render_template("semestres/eliminarCurso_semestre.html", title="Eliminar curso a semestre", form=form, semestre=semestre, curso=curso)


@semestres.route("/semestres/<int:semestre_id>/eliminarSemestre", methods=["GET", "POST"])
@login_required
def eliminar_semestre(semestre_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    semestre = Semestre.query.get_or_404(semestre_id)
    actividades = Actividad.query.filter_by(semestre_id=semestre.id).all()
    estudiantes = Estudiante.query.filter_by(semestre=semestre.id).all()
    form = EliminarSemestreForm()
    if form.validate_on_submit():
        for curso in semestre.cursos:
            semestre.cursos.remove(curso)
            semestreCurso = SemestreCursoHabilitados.query.filter_by(semestre_id=semestre.id, curso_id=curso.id).first()
            db.session.delete(semestreCurso)
            db.session.commit()
        for estudiante in estudiantes:
            semestre.estudiantes.remove(estudiante)
            db.session.delete(estudiante)
            db.session.commit()
        for actividad in actividades:
            grupos = Grupo.query.filter_by(actividad_id=actividad.id).all()
            for grupo in grupos:
                listaCalificaciones = Calificacion.query.filter_by(grupo_id=grupo.id).all()
                for calificacion in listaCalificaciones:
                    db.session.delete(calificacion)
                    db.session.commit()
                db.session.delete(grupo)
                db.session.commit()
            eliminarActividad(actividad.id)
            db.session.commit()
        db.session.delete(semestre)
        db.session.commit()
        flash(f"Semestre eliminado exitosamente", "success")
        return redirect(url_for("semestres.get_semestres"))
    return render_template("semestres/eliminar_semestre.html", title="Eliminar semestre", form=form, semestre=semestre)

@semestres.route("/semestres/<int:semestre_id>/<int:curso_id>/verLista", methods=["GET", "POST"])
@login_required
def verLista_semestre(semestre_id,curso_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    semestre = Semestre.query.get_or_404(semestre_id)
    return render_template("semestres/verLista_semestre.html", title="Ver lista estudiantes", semestre=semestre, curso_id=curso_id)

@semestres.route("/semestres/<int:semestre_id>/<int:curso_id>/cargarLista", methods=["GET", "POST"])
@login_required
def cargarListaEstudiantes_semestre(semestre_id,curso_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    semestre = Semestre.query.get_or_404(semestre_id)
    form = CargarListaEstudiantesForm()
    if form.validate_on_submit():
        if form.archivo.data:
            if request.method == 'POST':
                f = request.files['archivo']
                f.save(os.path.join(current_app.root_path, 'static/files', "ListaEstudiantes.xlsx"))
                analisis = analizarArchivoListaEstudiantes(semestre)
                if analisis:
                    flash(f"Lista de estudiantes importada exitosamente", "success")
                    return redirect(url_for("semestres.verLista_semestre", curso_id=curso_id, semestre_id=semestre_id))
                else:
                    flash(f"El archivo no pudo ser procesado porque no cumple con la estructura.", "danger")
                    return redirect(url_for("semestre.verLista_semestre", curso_id=curso_id, semestre_id=semestre_id))
    return render_template("semestres/cargarLista_semestre.html", title="Añadir estudiante", semestre=semestre, curso_id=curso_id, form=form)

def analizarArchivoListaEstudiantes(semestre):
    exito = True
    archivoExcel = load_workbook(current_app.root_path + '/static/files/ListaEstudiantes.xlsx')
    hoja = archivoExcel.active
    fila = 2
    final = False
    while not final:
        codigo = hoja.cell(row=fila, column=1).value
        apellidos = hoja.cell(row=fila, column=2).value
        nombres = hoja.cell(row=fila, column=3).value
        login = hoja.cell(row=fila, column=4).value
        magistral = hoja.cell(row=fila, column=5).value
        complementaria = hoja.cell(row=fila, column=6).value
        if None in (codigo, apellidos, nombres, login, magistral, complementaria):
            final = True
        else:
            if (Estudiante.query.filter_by(codigo=codigo, semestre=semestre.id).first() == None and Estudiante.query.filter_by(login=login, semestre=semestre.id).first() == None):
                estudiante = Estudiante(codigo=codigo, login=login, apellido=apellidos, nombre=nombres, magistral=magistral, complementaria=complementaria)
                semestre.estudiantes.append(estudiante)
                db.session.add(estudiante)
                db.session.commit()
                fila = fila + 1
            else:
                exito = False
                final = True
    return exito

@semestres.route("/semestres/<int:semestre_id>/<int:curso_id>/descargarLista", methods=["GET", "POST"])
@login_required
def descargarListaEstudiantes_semestre(semestre_id,curso_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    semestre = Semestre.query.get_or_404(semestre_id)
    wb = Workbook()
    dest_filename = 'Herramienta/static/files/ListaDeEstudiantesExportar.xlsx'
    hoja = wb.active
    hoja.cell(column=1, row=1, value="Código:")
    hoja.cell(column=2, row=1, value="Apellidos:")
    hoja.cell(column=3, row=1, value="Nombres:")
    hoja.cell(column=4, row=1, value="Login:")
    hoja.cell(column=5, row=1, value="Magistral:")
    hoja.cell(column=6, row=1, value="Complementaria:")

    fila = 2

    for estudiante in semestre.estudiantes:
        hoja.cell(column=1, row=fila, value=estudiante.codigo)
        hoja.cell(column=2, row=fila, value=estudiante.apellido)
        hoja.cell(column=3, row=fila, value=estudiante.nombre)
        hoja.cell(column=4, row=fila, value=estudiante.login)
        hoja.cell(column=5, row=fila, value=estudiante.magistral)
        hoja.cell(column=6, row=fila, value=estudiante.complementaria)
        fila = fila + 1

    wb.save(filename = dest_filename)

    form = DescargarListaEstudiantesForm()
    if form.validate_on_submit():
        return send_file('static/files/ListaDeEstudiantesExportar.xlsx',
                     mimetype='text/xlsx',
                     attachment_filename='ListaDeEstudiantesExportar.xlsx',
                     as_attachment=True)
    return render_template("semestres/descargarListaEstudiantes.html", title="Descargar lista de estudiantes", semestre=semestre, curso_id=curso_id, form=form, semestre_id=semestre_id)

@semestres.route("/semestres/<int:semestre_id>/<int:curso_id>/descargarFormatoLista", methods=["GET", "POST"])
@login_required
def descargarFormatoListaEstudiantes_semestre(semestre_id,curso_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    semestre = Semestre.query.get_or_404(semestre_id)
    wb = Workbook()
    dest_filename = 'Herramienta/static/files/FormatoListaDeEstudiantesExportar.xlsx'
    hoja = wb.active
    hoja.cell(column=1, row=1, value="Código:")
    hoja.cell(column=2, row=1, value="Apellidos:")
    hoja.cell(column=3, row=1, value="Nombres:")
    hoja.cell(column=4, row=1, value="Login:")
    hoja.cell(column=5, row=1, value="Magistral:")
    hoja.cell(column=6, row=1, value="Complementaria:")

    wb.save(filename = dest_filename)

    form = DescargarFormatoListaEstudiantesForm()
    if form.validate_on_submit():
        return send_file('static/files/FormatoListaDeEstudiantesExportar.xlsx',
                     mimetype='text/xlsx',
                     attachment_filename='FormatoListaDeEstudiantesExportar.xlsx',
                     as_attachment=True)
    return render_template("semestres/descargarFormatoListaEstudiantes.html", title="Descargar formato lista de estudiantes", semestre=semestre, curso_id=curso_id, form=form, semestre_id=semestre_id)

@semestres.route("/semestres/<int:semestre_id>/<int:curso_id>/<int:estudiante_id>/editarEstudiante", methods=["GET", "POST"])
@login_required
def editarEstudiante_semestre(semestre_id,curso_id,estudiante_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    estudiante = Estudiante.query.get_or_404(estudiante_id)
    form = EstudianteForm()
    if form.validate_on_submit():
        estudiante.nombre = form.nombres.data
        estudiante.apellido = form.apellidos.data
        estudiante.codigo = form.codigo.data
        estudiante.login = form.login.data
        estudiante.magistral = form.magistral.data
        estudiante.complementaria = form.complementaria.data
        db.session.commit()
        flash("Estudiante actualizado con éxito.", "success")
        return redirect(url_for("semestres.verLista_semestre", semestre_id=semestre_id, curso_id=curso_id))
    elif request.method == "GET":
        form.nombres.data = estudiante.nombre
        form.apellidos.data = estudiante.apellido
        form.codigo.data = estudiante.codigo
        form.login.data = estudiante.login
        form.magistral.data = estudiante.magistral
        form.complementaria.data = estudiante.complementaria
    return render_template("semestres/estudiante.html", title="Editar estudiante", form=form, legend="Editar estudiante", curso_id=curso_id, semestre_id=semestre_id)

@semestres.route("/semestres/<int:semestre_id>/<int:curso_id>/crearEstudiante", methods=["GET", "POST"])
@login_required
def crearEstudiante_semestre(semestre_id,curso_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    semestre = Semestre.query.get_or_404(semestre_id)
    form = EstudianteForm()
    estudiantes = semestre.estudiantes
    estudiantesJSON = []
    for estudiante in estudiantes:
        estudianteAnadir = {
            "codigo" : str(estudiante.codigo),
            "login" : estudiante.login
        }
        estudiantesJSON.append(estudianteAnadir)
    if form.validate_on_submit():
        existenteCodigo = Estudiante.query.filter_by(codigo=form.codigo.data, semestre=semestre.id).first()
        existenteLogin = Estudiante.query.filter_by(login=form.login.data, semestre=semestre.id).first()
        if existenteCodigo is not None:
            flash("Ya existe un estudiante con ese código en este semestre", "danger")
        elif existenteLogin is not None:
            flash("Ya existe un estudiante con ese login en este semestre", "danger")
        else:
            estudiante = Estudiante(codigo=form.codigo.data, login=form.login.data, apellido=form.apellidos.data, nombre=form.nombres.data, magistral=form.magistral.data, complementaria=form.complementaria.data)
            semestre.estudiantes.append(estudiante)
            db.session.add(estudiante)
            db.session.commit()
            flash("Estudiante creado con éxito.", "success")
        return redirect(url_for("semestres.verLista_semestre", semestre_id=semestre_id, curso_id=curso_id))
    return render_template("semestres/estudiante.html", title="Crear estudiante", form=form, legend="Crear estudiante", curso_id=curso_id, semestre_id=semestre_id, estudiantesJSON=estudiantesJSON)

@semestres.route("/semestres/<int:semestre_id>/<int:curso_id>/<int:estudiante_id>/eliminarEstudiante", methods=["GET", "POST"])
@login_required
def eliminar_estudiante(semestre_id,curso_id, estudiante_id):
    if not current_user.activado:
        return redirect(url_for("usuarios.activar_usuario"))
    user_id = current_user.get_id()
    user = Usuario.query.filter_by(id=user_id).first()
    if user.rol_id == 1:
        abort(403)
    estudiante = Estudiante.query.get_or_404(estudiante_id)
    semestre = Semestre.query.get_or_404(semestre_id)
    curso = Curso.query.get_or_404(curso_id)
    form = EliminarEstudianteForm()
    if form.validate_on_submit():
        semestre.estudiantes.remove(estudiante)
        db.session.delete(estudiante)
        db.session.commit()
        flash(f"Estudiante eliminado exitosamente", "success")
        return redirect(url_for("semestres.verLista_semestre", semestre_id=semestre_id, curso_id=curso_id))
    return render_template("semestres/eliminar_estudiante.html", title="Eliminar estudiante", form=form, legend="Eliminar estudiante", curso=curso, semestre=semestre, estudiante=estudiante)

def eliminarActividad(actividad_id):
    actividad = Actividad.query.get_or_404(actividad_id)

    for punto in actividad.puntos:
        #-------------------------------------------------------------------
        for inciso in punto.incisos:
            #-------------------------------------------------------------------
            for criterio in inciso.criterios:
                #-------------------------------------------------------------------
                for subcriterio in criterio.subcriterios:
                    #-------------------------------------------------------------------
                    for variacion in subcriterio.variaciones:
                        db.session.delete(variacion)
                        db.session.commit()
                    #-------------------------------------------------------------------
                    db.session.delete(subcriterio)
                    db.session.commit()
                #-------------------------------------------------------------------
                db.session.delete(criterio)
                db.session.commit()
            #-------------------------------------------------------------------
            db.session.delete(inciso)
            db.session.commit()
        #-------------------------------------------------------------------
        db.session.delete(punto)
        db.session.commit()
    
    db.session.delete(actividad)
    db.session.commit()